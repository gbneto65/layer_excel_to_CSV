# Layer Project V2
# database read and process
#
# Layers performance project - V2
# module: create a CSV database from Excel file (original data)
# 
# to be prepare the database  for the layers performance system
# 
# create a new CSV file everytime the database from de layers was updated or new layer genetics was inputed.
# as the CSV file is the file that will be load into the system.
# do not create any new sheet without the same strucutre of the others.
# the sheet_name should be the name of the genetic line of the layers
#
# Feb. 23th 2021

# know issues: it seems the Openpyxl module is corrupting the xlsx files. 


import pandas as pd
import os
import datetime

from openpyxl import load_workbook
# setup_parameters.py file contain the variables and functions
from setup_parameters import soft_version, \
    working_directory, excel_file_name, excel_col_import, csv_file_name, sound_error, intro_text, error_exit, \
    datetime_now


class DatabaseExcelRead():
    # read the database and
    def __init__(self):
        self.working_directory = working_directory
        self.excel_file_name = excel_file_name
        self.excel_col_import = excel_col_import
        self.csv_file_name = excel_file_name
        self.soft_version = soft_version

    # test if excel file exit on the current path
    # def date_time_now(self):
    #     return self.date_time_now

    def test_if_excel_exist(self):
        if os.path.isfile(self.excel_file_name):
            print('Excel File Exist')
            return True
        else:
            print('ERROR: Excel File does not exist')
            sound_error()  # call a function for sound error
            return False

    # read the spreadsheets names
    def get_sheet_names(self):
        # return self.excel_file_name
        # self.wb = load_workbook(self.excel_file_name, read_only = True, keep_links = False)
        self.wb = load_workbook(self.excel_file_name, read_only=True)
        self.ws = self.wb.sheetnames
        # print(len(self.ws))
        return self.ws

    def read_excel_file(self):
        dta = []
        for i in range(len(self.ws)):
            dta.append(i)

        for i in range(len(self.ws)):
            dta[i] = pd.read_excel(self.excel_file_name, sheet_name=self.ws[i], usecols=self.excel_col_import)
        self.layers_dta = pd.concat(dta, axis=0)
        print('Database concatenated with success')
        return self.layers_dta

    def csv_verify_if_exist(self):

        if os.path.isfile(self.csv_file_name):
            print(f'CSV File "{self.csv_file_name}" exist')
            return True
        else:
            print('ERROR: CSV File does not exist')
            sound_error()  # call a function for sound error
            return False


